import os
import re
import zipfile

import pandas
import openpyxl

from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font

from mylib.gservice import GService

def download_photo(link_photo, service_drive, target_file_name, dir_path):
    if not (pandas.isna(link_photo) or link_photo == ""):
        file_id = link_photo.split("id=")[1]
        file_info = service_drive.files().get(fileId=file_id, fields="name, fileExtension").execute()
        file_name = "{0}.{1}".format(target_file_name, file_info["fileExtension"])
        file_path = os.path.join(dir_path, file_name)
        if not os.path.exists(file_path):
            gservice.download_media(service_drive, file_id, file_path)
        return file_name
    return None

def remove_illegal_symbol(unfiltred_string):
    filter_pattern = r"[0-9a-zA-Zа-яА-Я_\+-=, \.\(\)]+"
    filtred_string = unfiltred_string.replace("/", "_")
    filtred_string_parts = re.findall(filter_pattern, filtred_string)
    filtred_string = "".join(filtred_string_parts)
    return filtred_string

if __name__ == "__main__":
    token = os.path.join("token", "python-trade-points-audit.json")
    SPREADSHEET_ID = "1pQE5VS2Vf69oHbo6P-Xs2BVIHlBDCrT_jfjv7smg-f4"
    SCOPES = [
        "https://www.googleapis.com/auth/spreadsheets.readonly",
        "https://www.googleapis.com/auth/drive.readonly"
    ]

    AUDIT_CODE = "YTHNI09"
    AUDIT_NAME = "Алматы 2020-09-02 - 2020-09-04 test"

    dir_root = "\\\\aw.com\\cloud\\REPORTS\\АУДИТ\\Almaty"
    dir_audit_photos = os.path.join(dir_root, AUDIT_NAME)
    if not os.path.exists(dir_audit_photos):
        os.mkdir(dir_audit_photos)

    print("Выгрузка аудита: %s" % AUDIT_NAME)

    # temp_dir = ".\\temp"
    # if not os.path.exists(temp_dir):
    #     os.mkdir(temp_dir)

    gservice = GService(SCOPES)
    gservice.auth(token)
    spreadsheets = gservice.get_spreadsheets()
    service_gdrive = gservice.get_service_drive()

    sheets = spreadsheets.get(spreadsheetId=SPREADSHEET_ID).execute()
    range_name = sheets["sheets"][0]["properties"]["title"]
    sheet_rows = spreadsheets.values().get(spreadsheetId=SPREADSHEET_ID, range=range_name).execute()
    values = sheet_rows.get("values", [])
    if len(values) <= 1:  # первая строка - заголовок
        raise Exception("No data found!")

    header = values[0]
    df_data = pandas.DataFrame(values)
    df_data.columns = header
    df_data = df_data.drop(0, axis="index")

    col_names_code ={
        "{1}": "AuditorName",
        "{3}": "AuditCode",
        "{5}": "SalesPointTitle",
        "{6}": "Address",
        "{7}": "SalesChannel",
        "{42}": "Comment",

        "{4}": "link_FacadePhoto",
        "{9}": "link_ShelfPhoto_general_vodka",
        "{10}": "link_ShelfPhoto_closeup_vodka",
        "{35}": "link_ShelfPhoto_general_cognac",
        "{36}": "link_ShelfPhoto_closeup_cognac"
    }

    columns_new_names = {
        "Отметка времени": "Date"
    }

    for column_name in df_data.columns:
        for column_code in col_names_code:
            if str(column_code) in str(column_name):
                columns_new_names[column_name] = col_names_code[column_code]

    df_data = df_data.rename(columns=columns_new_names)

    df_data = df_data[df_data["AuditCode"] == AUDIT_CODE]
    df_data = df_data[df_data["Comment"].notna()]
    df_data["AuditName"] = AUDIT_NAME

    # удаление дублей
    df_data['row_key'] = df_data['AuditCode'] + '-' + df_data['SalesPointTitle'] + '-' + df_data['Address']
    df_data_unique = []
    for row_key in df_data['row_key'].unique():
        df_data_unique.append(df_data[df_data['row_key'] == row_key].iloc[0])
    df_data = pandas.DataFrame(df_data_unique)

    # arcfiles_list = []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "sheet1"

    header_row = 1
    first_row = 2
    first_col = 1

    # columns width
    ws.column_dimensions[get_column_letter(first_col + 0)].width = 35
    ws.column_dimensions[get_column_letter(first_col + 1)].width = 20
    ws.column_dimensions[get_column_letter(first_col + 2)].width = 12
    ws.column_dimensions[get_column_letter(first_col + 3)].width = 20
    ws.column_dimensions[get_column_letter(first_col + 4)].width = 40
    ws.column_dimensions[get_column_letter(first_col + 5)].width = 50
    ws.column_dimensions[get_column_letter(first_col + 6)].width = 15
    ws.column_dimensions[get_column_letter(first_col + 7)].width = 15
    ws.column_dimensions[get_column_letter(first_col + 8)].width = 15
    ws.column_dimensions[get_column_letter(first_col + 9)].width = 15
    ws.column_dimensions[get_column_letter(first_col + 10)].width = 15
    # cell styles
    # alignment_left_top_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
    alignment_header = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alignment_cell = Alignment(horizontal='left', vertical='top', wrap_text=True)
    # cell border
    side_thin = Side(style="thin", color="000000")
    border_all = Border(left=side_thin, right=side_thin, top=side_thin, bottom=side_thin)
    # cell font
    font_bold = Font(bold=True)
    # header
    ws.cell(header_row, first_col + 0).value = "Аудит"
    ws.cell(header_row, first_col + 1).value = "Аудитор"
    ws.cell(header_row, first_col + 2).value = "Категория точки"
    ws.cell(header_row, first_col + 3).value = "Точка"
    ws.cell(header_row, first_col + 4).value = "Адрес"
    ws.cell(header_row, first_col + 5).value = "Комментарий"
    ws.cell(header_row, first_col + 6).value = "Фото фасада"
    ws.cell(header_row, first_col + 7).value = "Общее фото водочной полки"
    ws.cell(header_row, first_col + 8).value = "Фото водочной полки крупным планом"
    ws.cell(header_row, first_col + 9).value = "Общее фото коньячной полки"
    ws.cell(header_row, first_col + 10).value = "Фото коньячной полки крупным планом"

    for row in ws["A1:{0}1".format(get_column_letter(first_col + 10))]:
        for cell in row:
            cell.alignment = alignment_header
            cell.border = border_all
            cell.font = font_bold

    i = first_row
    for row in df_data.iterrows():
        row_data = row[1]
        ws.cell(i, first_col + 0).value = row_data["AuditName"]
        ws.cell(i, first_col + 1).value = row_data["AuditorName"]
        ws.cell(i, first_col + 2).value = row_data["SalesChannel"]
        ws.cell(i, first_col + 3).value = row_data["SalesPointTitle"]
        ws.cell(i, first_col + 4).value = row_data["Address"]
        ws.cell(i, first_col + 5).value = row_data["Comment"]
        # photo name parts
        filtred_sales_point = remove_illegal_symbol(row_data["SalesPointTitle"])
        filtred_address = remove_illegal_symbol(row_data["Address"])
        # download photo
        dir_sales_point_name = filtred_sales_point.strip() + "_" + filtred_address.strip()
        sales_point_dir = os.path.join(dir_audit_photos, dir_sales_point_name)
        if not os.path.exists(sales_point_dir):
            os.mkdir(sales_point_dir)
        photo_name = "facade_photo_" + dir_sales_point_name
        file_name_facade_photo = download_photo(row_data["link_FacadePhoto"], service_gdrive, photo_name, sales_point_dir)
        photo_name = "vodka_general" + dir_sales_point_name
        file_name_vodka_general = download_photo(row_data["link_ShelfPhoto_general_vodka"], service_gdrive, photo_name, sales_point_dir)
        photo_name = "vodka_closeup" + dir_sales_point_name
        file_name_vodka_closeup = download_photo(row_data["link_ShelfPhoto_closeup_vodka"], service_gdrive, photo_name, sales_point_dir)
        photo_name = "cognac_general" + dir_sales_point_name
        file_name_cognac_general = download_photo(row_data["link_ShelfPhoto_general_cognac"], service_gdrive, photo_name, sales_point_dir)
        photo_name = "cognac_closeup" + dir_sales_point_name
        file_name_cognac_closeup = download_photo(row_data["link_ShelfPhoto_closeup_cognac"], service_gdrive, photo_name, sales_point_dir)
        # cell formula
        if file_name_facade_photo is not None:
            photo_link = os.path.join(AUDIT_NAME, dir_sales_point_name, file_name_facade_photo)
            ws.cell(i, first_col + 6).value = "=HYPERLINK(\"{0}\", \"Фото фасада\")".format(photo_link)
            # arcfile = {
            #     "file_name": os.path.join(temp_dir, dir_name, file_name_facade_photo),
            #     "arcname": os.path.join(AUDIT_NAME, dir_name, file_name_facade_photo)
            # }
            # arcfiles_list.append(arcfile)
        if file_name_vodka_general is not None:
            photo_link = os.path.join(AUDIT_NAME, dir_sales_point_name, file_name_vodka_general)
            ws.cell(i, first_col + 7).value = "=HYPERLINK(\"{0}\", \"Общее фото водочной полки\")".format(photo_link)
            # arcfile = {
            #     "file_name": os.path.join(temp_dir, dir_name, file_name_vodka_general),
            #     "arcname": os.path.join(AUDIT_NAME, dir_name, file_name_vodka_general)
            # }
            # arcfiles_list.append(arcfile)
        if file_name_vodka_closeup is not None:
            photo_link = os.path.join(AUDIT_NAME, dir_sales_point_name, file_name_vodka_closeup)
            ws.cell(i, first_col + 8).value = "=HYPERLINK(\"{0}\", \"Фото водочной полки крупным планом\")".format(photo_link)
            # arcfile = {
            #     "file_name": os.path.join(temp_dir, dir_name, file_name_vodka_closeup),
            #     "arcname": os.path.join(AUDIT_NAME, dir_name, file_name_vodka_closeup)
            # }
            # arcfiles_list.append(arcfile)
        if file_name_cognac_general is not None:
            photo_link = os.path.join(AUDIT_NAME, dir_sales_point_name, file_name_cognac_general)
            ws.cell(i, first_col + 9).value = "=HYPERLINK(\"{0}\", \"Общее фото коньячной полки\")".format(photo_link)
            # arcfile = {
            #     "file_name": os.path.join(temp_dir, dir_name, file_name_cognac_general),
            #     "arcname": os.path.join(AUDIT_NAME, dir_name, file_name_cognac_general)
            # }
            # arcfiles_list.append(arcfile)
        if file_name_cognac_closeup is not None:
            photo_link = os.path.join(AUDIT_NAME, dir_sales_point_name, file_name_cognac_closeup)
            ws.cell(i, first_col + 10).value = "=HYPERLINK(\"{0}\", \"Фото коньячной полки крупным планом\")".format(photo_link)
            # arcfile = {
            #     "file_name": os.path.join(temp_dir, dir_name, file_name_cognac_closeup),
            #     "arcname": os.path.join(AUDIT_NAME, dir_name, file_name_cognac_closeup)
            # }
            # arcfiles_list.append(arcfile)
        print(filtred_sales_point)
        i += 1

    # text cells style
    cells_range = "A{0}:{1}{2}".format(first_row, get_column_letter(6), i-1)
    for row in ws[cells_range]:
        for cell in row:
            cell.alignment = alignment_cell
            cell.border = border_all
    # hyperlink cells style
    cells_range = "{0}{1}:{2}{3}".format(get_column_letter(7), first_row, get_column_letter(first_col + 10), i-1)
    for row in ws[cells_range]:
        for cell in row:
            cell.style = 'Hyperlink'
            cell.alignment = alignment_cell
            cell.border = border_all
    excel_file_name = AUDIT_NAME + ".xlsx"
    wb.save(os.path.join(dir_root, excel_file_name))

    # zip_file = zipfile.ZipFile(AUDIT_NAME + ".zip", mode="w")
    # for file_names in arcfiles_list:
    #     zip_file.write(file_names["file_name"], file_names["arcname"])
    # zip_file.write(os.path.join(temp_dir, excel_file_name), os.path.join(AUDIT_NAME, excel_file_name))
    # zip_file.close()
